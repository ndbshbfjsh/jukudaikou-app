from flask import Flask, render_template, request, redirect, send_file, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pywebpush import webpush, WebPushException
import os
import io
import json
import calendar

calendar.setfirstweekday(calendar.SUNDAY)

app = Flask(__name__)
app.secret_key = "anything-secret"

db_url = os.getenv("DATABASE_URL")
if db_url:
    db_url = db_url.strip()
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url or "sqlite:///instance/database.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 280
}

db = SQLAlchemy(app)


class Data(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    branch = db.Column(db.String(100))
    date = db.Column(db.String(20))
    period = db.Column(db.String(20))
    time = db.Column(db.String(50))
    teacher = db.Column(db.String(100))
    status = db.Column(db.String(20))
    sub_teacher = db.Column(db.String(100))
    note = db.Column(db.String(200))


class PushSubscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True)
    endpoint = db.Column(db.Text, unique=True)
    subscription_json = db.Column(db.Text)


class NotificationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True)


with app.app_context():
    db.create_all()
    try:
        db.session.execute(db.text("DROP TABLE IF EXISTS push_subscription"))
        db.session.commit()
    except Exception as e:
        print("DROP push_subscription error:", e)

    db.create_all()

SUNDAY_PERIOD_TIMES = {
    "1": "1限　9:30 ～ 10:50",
    "2": "2限　11:00 ～ 12:20",
    "3": "3限　13:20 ～ 14:40",
    "4": "4限　15:00 ～ 16:20",
    "5": "5限　16:30 ～ 17:50",
    "6": "6限　18:05 ～ 19:25",
    "7": "7限　19:35 ～ 20:55",
}

SATURDAY_PERIOD_TIMES = {
    "1": "1限　9:30 ～ 10:50",
    "1-1": "1限 1/2　9:20 ～ 10:05",
    "1-2": "1限 2/2　10:05 ～ 10:50",
    "2": "2限　11:00 ～ 12:20",
    "2-1": "2限 1/2　10:50 ～ 11:35",
    "2-2": "2限 2/2　11:35 ～ 12:20",
    "3s": "3限（小学生集団）13:20 ～ 14:20",
    "3": "3限　13:20 ～ 14:40",
    "3-1": "3限 1/2　13:10 ～ 13:55",
    "3-2": "3限 2/2　13:55 ～ 14:40",
    "4": "4限　14:50 ～ 16:10",
    "4-1": "4限 1/2　14:50 ～ 15:35",
    "4-2": "4限 2/2　15:35 ～ 16:20",
    "5": "5限　16:30 ～ 17:50",
    "6": "6限　18:05 ～ 19:25",
    "7": "7限　19:35 ～ 20:55",
}

WEEKDAY_PERIOD_TIMES = {
    "1-1": "1限 1/2　9:20 ～ 10:05",
    "1-2": "1限 2/2　10:05 ～ 10:50",
    "2": "2限　11:00 ～ 12:20",
    "2-1": "2限 1/2　10:50 ～ 11:35",
    "2-2": "2限 2/2　11:35 ～ 12:20",
    "3s": "3限（小学生集団）13:20 ～ 14:20",
    "3": "3限　13:20 ～ 14:40",
    "3-1": "3限 1/2　13:10 ～ 13:55",
    "3-2": "3限 2/2　13:55 ～ 14:40",
    "4": "4限　14:50 ～ 16:10",
    "4-1": "4限 1/2　15:25 ～ 16:10",
    "4-2": "4限 2/2　16:10 ～ 16:55",
    "5s": "5限（小学生集団）17:15 ～ 18:15",
    "5": "5限　17:15 ～ 18:35",
    "5-1": "5限 1/2　17:05 ～ 17:50",
    "5-2": "5限 2/2　17:50 ～ 18:35",
    "6": "6限　18:45 ～ 20:05",
    "7": "7限　20:15 ～ 21:35",
}
SUMMER_PERIOD_TIMES = {
    "1": "1限　9:30 ～ 10:50",
    "2": "2限　11:00 ～ 12:20",
    "3": "3限　13:30 ～ 14:50",
    "4": "4限　15:00 ～ 16:20",
    "5": "5限　16:30 ～ 17:50",
    "6": "6限　18:05 ～ 19:25",
    "7": "7限　19:35 ～ 20:55",
}

def get_period_times(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")

    summer_ranges = [
        ("2026-07-26", "2026-07-30"),
        ("2026-08-01", "2026-08-05"),
        ("2026-08-17", "2026-08-21"),
        ("2026-08-23", "2026-08-27"),
    ]

    for start, end in summer_ranges:
        if start <= date_str <= end:
            return SUMMER_PERIOD_TIMES

    if d.weekday() == 6:
        return SUNDAY_PERIOD_TIMES
    elif d.weekday() == 5:
        return SATURDAY_PERIOD_TIMES

    return WEEKDAY_PERIOD_TIMES


def calc_minutes(time_text):
    try:
        time_part = time_text.split("　")[-1]
        if "～" not in time_part:
            return ""
        start, end = time_part.split("～")
        start_dt = datetime.strptime(start.strip(), "%H:%M")
        end_dt = datetime.strptime(end.strip(), "%H:%M")
        return int((end_dt - start_dt).total_seconds() / 60)
    except:
        return ""


def send_push_notification(title, body):
    public_key = os.getenv("VAPID_PUBLIC_KEY")
    private_key = os.getenv("VAPID_PRIVATE_KEY")

    print("PUSH START")
    print("PUBLIC KEY exists:", bool(public_key))
    print("PRIVATE KEY exists:", bool(private_key))

    if not public_key or not private_key:
        print("VAPID KEY MISSING")
        return

    subscriptions = PushSubscription.query.all()
    print("SUB COUNT:", len(subscriptions))

    for sub in subscriptions:
        try:
            print("SEND TO:", sub.endpoint[:50])

            webpush(
                subscription_info=json.loads(sub.subscription_json),
                data=json.dumps({
                    "title": title,
                    "body": body
                }),
                vapid_private_key=private_key,
                vapid_claims={
                    "sub": "mailto:haozaisangu5@gmail.com"
                }
            )

            print("PUSH SUCCESS")

        except WebPushException as e:
            print("PUSH WEB ERROR:", repr(e))
            print("RESPONSE:", getattr(e, "response", None))

        except Exception as e:
            print("PUSH ERROR:", repr(e))
JST = timezone(timedelta(hours=9))
def delete_old_records():
    today = datetime.now(JST).date()

    # 毎月10日より前は削除しない
    if today.day < 10:
        return

    # 前月を計算
    if today.month == 1:
        target_year = today.year - 1
        target_month = 12
    else:
        target_year = today.year
        target_month = today.month - 1

    start_date = f"{target_year:04d}-{target_month:02d}-01"

    # 前月の最終日
    last_day = calendar.monthrange(target_year, target_month)[1]
    end_date = f"{target_year:04d}-{target_month:02d}-{last_day:02d}"

    old_records = Data.query.filter(
        Data.date >= start_date,
        Data.date <= end_date
    ).all()

    for r in old_records:
        db.session.delete(r)

    db.session.commit()
def notify_today_pending():
    today = datetime.now(JST).strftime("%Y-%m-%d")
    log_key = f"today_pending_{today}"

    if NotificationLog.query.filter_by(key=log_key).first():
        return

    records = Data.query.filter_by(date=today, status="未定").all()

    if not records:
        return

    lines = []
    for r in records:
        lines.append(f"{r.period}限　担当：{r.teacher}")

    body = f"{today}\n未定の代講があります\n" + "\n".join(lines)

    send_push_notification(
        "⚠️ 本日まだ未定の代講があります",
        body
    )

    log = NotificationLog(key=log_key)
    db.session.add(log)
    db.session.commit()

@app.route("/vapid_public_key")
def vapid_public_key():
    return jsonify({
        "publicKey": os.getenv("VAPID_PUBLIC_KEY", "")
    })


@app.route("/subscribe", methods=["POST"])
def subscribe():
    sub_data = request.get_json()
    endpoint = sub_data.get("endpoint")

    existing = PushSubscription.query.filter_by(endpoint=endpoint).first()
    if not existing:
        new_sub = PushSubscription(
            endpoint=endpoint,
            subscription_json=json.dumps(sub_data)
        )
        db.session.add(new_sub)
        db.session.commit()

    return jsonify({"status": "ok"})
@app.route("/unsubscribe", methods=["POST"])
def unsubscribe():
    data = request.get_json()
    endpoint = data.get("endpoint")

    sub = PushSubscription.query.filter_by(endpoint=endpoint).first()

    if sub:
        db.session.delete(sub)
        db.session.commit()

    return jsonify({"status": "ok"})

@app.route("/")
def index():
    delete_old_records()
    notify_today_pending()
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    today_dt = datetime.today()

    if not year or not month:
        year = today_dt.year
        month = today_dt.month

    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    cal = calendar.monthcalendar(year, month)

    start_date = f"{year:04d}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    end_date = f"{year:04d}-{month:02d}-{last_day:02d}"

    records = Data.query.filter(
        Data.date >= start_date,
        Data.date <= end_date
    ).all()

    teacher_status_by_date = defaultdict(dict)

    for r in records:
        if not r.teacher:
            continue
        teacher_status_by_date[r.date][r.teacher] = r.status

    count_by_date = {}

    for date_key, teachers in teacher_status_by_date.items():
        blue = 0
        red = 0
        yellow = 0

        for teacher, status in teachers.items():
            if status == "決定":
                blue += 1
            elif status == "未定":
                red += 1
            else:
                yellow += 1

        count_by_date[date_key] = {
            "blue": blue,
            "red": red,
            "yellow": yellow
        }

    return render_template(
        "index.html",
        year=year,
        month=month,
        cal=cal,
        count_by_date=count_by_date,
        today=today_dt.strftime("%Y-%m-%d")
    )


@app.route("/day/<date>", methods=["GET", "POST"])
def day(date):
    period_times = get_period_times(date)

    if request.method == "POST":
        action = request.form.get("action")

        if action == "bulk_delete":
            selected_ids = request.form.getlist("selected_ids")
            for id in selected_ids:
                data = Data.query.get(id)
                if data:
                    db.session.delete(data)
            db.session.commit()
            
            return redirect(f"/day/{date}")

        if action == "bulk_update":
            selected_ids = request.form.getlist("selected_ids")
            bulk_teacher = request.form.get("bulk_teacher")
            bulk_status = request.form.get("bulk_status")
            bulk_sub_teacher = request.form.get("bulk_sub_teacher")
            bulk_note = request.form.get("bulk_note")

            for id in selected_ids:
                data = Data.query.get(id)
                if data:
                    if bulk_teacher:
                        data.teacher = bulk_teacher
                    if bulk_status:
                        data.status = bulk_status
                        if bulk_status == "未定":
                            data.sub_teacher = ""
                    if bulk_sub_teacher:
                        data.sub_teacher = bulk_sub_teacher
                    if bulk_note:
                        data.note = bulk_note

            db.session.commit()
            send_push_notification("塾代講管理", f"{date} の代講を更新しました")
            return redirect(f"/day/{date}")

        branch = request.form.get("branch")
        teacher = request.form.get("teacher")
        status = request.form.get("status")
        sub_teacher = request.form.get("sub_teacher")
        note = request.form.get("note")
        periods = request.form.getlist("period")

        if status == "未定":
            sub_teacher = ""

        for p in periods:
            new_data = Data(
                branch=branch,
                date=date,
                period=p,
                time=period_times.get(p, ""),
                teacher=teacher,
                status=status,
                sub_teacher=sub_teacher,
                note=note
            )
            db.session.add(new_data)

        db.session.commit()

        send_push_notification(
            "📢 新しい代講が登録されました",
            f"{date}\n担当：{teacher}\n状態：{status}"
        )

        return redirect(f"/day/{date}")

    records = Data.query.filter_by(date=date).order_by(Data.period).all()

    records_by_period = defaultdict(list)
    for r in records:
        records_by_period[r.period].append(r)

    return render_template(
        "day.html",
        date=date,
        period_times=period_times,
        records_by_period=records_by_period,
        records=records,
        calc_minutes=calc_minutes,
        year=int(date[:4]),
        month=int(date[5:7])
)


@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):
    data = Data.query.get_or_404(id)
    period_times = get_period_times(data.date)

    if request.method == "POST":
        data.teacher = request.form.get("teacher")
        data.status = request.form.get("status")
        data.sub_teacher = request.form.get("sub_teacher")
        data.note = request.form.get("note")

        if data.status == "未定":
            data.sub_teacher = ""

        db.session.commit()

        send_push_notification(
            "塾代講管理",
            f"{data.date} の代講が更新されました\n担当：{data.teacher}\n状態：{data.status}"
        )

        return redirect(f"/day/{data.date}")

    return render_template(
        "edit.html",
        data=data,
        period_times=period_times
    )


@app.route("/delete/<int:id>")
def delete(id):
    data = Data.query.get_or_404(id)
    date = data.date
    db.session.delete(data)
    db.session.commit()

    

    return redirect(f"/day/{date}")


@app.route("/export")
def export():
    records = Data.query.order_by(Data.date, Data.period).all()

    wb = Workbook()
    ws = wb.active
    ws.append(["日付", "校舎", "限", "時間", "先生", "状態", "代講先生", "備考"])

    for r in records:
        ws.append([
            r.date,
            r.branch,
            r.period,
            r.time,
            r.teacher,
            r.status,
            r.sub_teacher,
            r.note
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="daikou.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/check_today_pending")
def check_today_pending():
    notify_today_pending()
    return "checked today pending"
@app.route("/clear_subscriptions")
def clear_subscriptions():
    PushSubscription.query.delete()
    db.session.commit()
    return "subscriptions cleared"
if __name__ == "__main__":
    app.run(debug=True)
@app.route("/help")
def help():
    return render_template("help.html")